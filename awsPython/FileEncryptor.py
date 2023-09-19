from cryptography.fernet import Fernet
from openpyxl import load_workbook


class Encryptor:
    def __init__(self):
        pass

    @staticmethod
    def key_create(self):
        key = Fernet.generate_key()
        return key

    def key_write(self, key, key_name):
        with open(key_name, 'wb') as mykey:
            mykey.write(key)

    def key_load(self, key_name):
        with open(key_name, 'rb') as mykey:
            key = mykey.read()
        return key

    def file_encrypt(self, key, original_file, encrypted_file):
        fernet = Fernet(key)
        with open(original_file, 'rb') as file:
            original = file.read()
            encrypted = fernet.encrypt(original)
        with open(encrypted_file, 'wb') as file:
            file.write(encrypted)

    def file_decrypt(self, key, encrypted_file, decrypted_file):
        fernet = Fernet(key)

        with open(encrypted_file, 'rb') as file:
            encrypted = file.read()
            decrypted = fernet.decrypt(encrypted)

        with open(decrypted_file, 'wb') as file:
            file.write(decrypted)

    def excel_encrypt(self, key, original_file, encrypted_file):
        f = Fernet(key)

        wb = load_workbook(filename=original_file)
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = f.encrypt(str(cell.value).encode())

        wb.save(filename=encrypted_file)

    def excel_decrypt(self, key, encrypted_file, decrypted_file):
        f = Fernet(key)

        wb = load_workbook(filename=encrypted_file)
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = f.decrypt(cell.value).decode()

        wb.save(filename=decrypted_file)
